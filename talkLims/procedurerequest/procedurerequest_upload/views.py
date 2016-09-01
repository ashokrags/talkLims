__author__ = "Ashok Ragavendran"
__license__ = "GPL"
__version__ = "1.0.1"
__email__ = "ashok.ragavendran@gmail.com"
__maintainer__ = "Ashok Ragavendran"
__status__ = "Production"

from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from jinja2 import FileSystemLoader
from openpyxl import load_workbook

from talkLims.models import TempPullRequest, RequestInfo, ProjectInfo
from talkLims.procedurerequest.procedurerequest_upload.procedurerequpload import ProcedureRequestUploadForm
from talkLims.utils.jinja2config import environment


@csrf_exempt
def initiate_request(request):
    # if this is a POST request we need to process the form data
    env = environment(loader=FileSystemLoader('/Users/ashok/PycharmProjects/talkLims/templates'))
    if request.method == 'POST':
        print request.FILES.keys()
        # create a form instance and populate it with data from the request:
        outForm = ProcedureRequestUploadForm(request.POST)

        # check whether it's valid:
        print "form loaded"
        if outForm.is_valid():
            print outForm.cleaned_data
        print "form validates"
        update_db(request.FILES['manifest_file'])

        env = environment(loader=FileSystemLoader('/Users/ashok/PycharmProjects/talkLims/templates'))
        template = env.get_template('base.html')
        return HttpResponse(template.render())
    else:
        myForm = ProcedureRequestUploadForm()
        env = environment(loader=FileSystemLoader('/Users/ashok/PycharmProjects/talkLims/templates'))
        template = env.get_template('procedure_request/procedure_request_upload_init.html')
        return HttpResponse(template.render(form=myForm))


def update_db(xls):
    update_db_project_info(xls)
    update_db_request_info(xls)
    update_db_temp(xls)
    return


def update_db_project_info(xls_file):
    wb = load_workbook(xls_file)
    ws = wb.active
    proj_data = dict()
    proj_data['project_name'] = ws['D5'].value
    proj_data['project_id'] = ws['D6'].value
    proj_data['project_driver'] = ws['D7'].value
    proj_data['driver_email'] = ws['D8'].value
    proj_data['protocol_type'] = ws['D9'].value
    proj_data['request_id'] = ws['D10'].value

    proj = ProjectInfo()

    p = [x.values()[0] for x in ProjectInfo.objects.values('request_id').values('request_id')]
    if len(p) > 0 and proj_data['request_id'] in p:
        print "Request ID exists in Project Table"
        return
    else:
        for (key, value) in proj_data.items():
            setattr(proj, key, value)
        proj.save()

    return


def update_db_request_info(xls_file):
    wb = load_workbook(xls_file)
    ws = wb.active
    req_tab_data = dict()
    req_tab_data['protocol_type'] = ws['D9'].value
    req_tab_data['request_id'] = ws['D10'].value
    req_tab_data['request_archivist'] = ws['D11'].value
    req_tab_data['requested_by'] = ws['D12'].value
    req_tab_data['request_date'] = ws['D13'].value
    req_tab_data['expected_end_date'] = ws['D14'].value
    req_tab_data['quote_number'] = ws['D15'].value
    req_tab_data['number_of_samples'] = ws['D16'].value
    req_tab_data['number_of_reads'] = ws['D17'].value
    req_tab_data['amount_required'] = ws['D18'].value
    req_tab_data['volume_required'] = ws['D19'].value

    r = RequestInfo()

    r1 = [x.values()[0] for x in RequestInfo.objects.values('request_id')]

    if len(r1) > 0 and req_tab_data['request_id'] in r1:
        print "Request ID exists in Request Table"
        return
    else:
        for (key, value) in req_tab_data.items():
            setattr(r, key, value)
        r.save()
    return


def update_db_temp(xls_file):
    wb = load_workbook(xls_file)
    ws = wb.active

    req_id = ws['D10'].value

    r1 = [x.values()[0] for x in TempPullRequest.objects.values('request_id')]

    if len(r1) > 0 and req_id in r1:
        print "Request ID exists in temp Table"
        return
    else:
        print "uploading temp Table"

    xls_keys = ['scan_id', 'sample_name', 'box_name', 'location_jane', 'shelf_jane',
                'rack_jane', 'slot_jane', 'grid', 'amount_jane', 'vol_jane', 'comments_from_request']

    for row in ws.iter_rows('A34:K1000'):
        p = TempPullRequest()

        ##upload_dat = {k:v for k,v in data.items()}
        upload_dat = {'request_id': req_id}

        samples = []
        cell_first = True
        for cell in row:
            if cell.value is not None:
                samples.append(cell.value)
            else:
                cell_first = False
                break
        if not cell_first:
            break
        else:
            sample_dat = dict(zip(xls_keys, samples))
            for k, v in sample_dat.items():
                upload_dat[k] = v
            for (key, value) in upload_dat.items():
                setattr(p, key, value)
            p.save()
            ##print upload_dat
    return
