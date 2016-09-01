__author__ = "Ashok Ragavendran"
__license__ = "GPL"
__version__ = "1.0.1"
__email__ = "ashok.ragavendran@gmail.com"
__maintainer__ = "Ashok Ragavendran"
__status__ = "Production"

import StringIO
import datetime

from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt

# from django.shortcuts import render

from jinja2 import FileSystemLoader
from openpyxl import load_workbook

from talkLims.settings import EXCELFILES_FOLDER, TEMPLATE_DIRS
from talkLims.utils.jinja2config import environment
from talkLims.models import ProjectInfo
from talkLims.procedurerequest.procedurerequest_download.procedurereqdownload import ProcedureRequestDownloadForm


@csrf_exempt
def initiate_request(request):
    # if this is a POST request we need to process the form data
    env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
    if request.method == 'POST':
        # create a form instance and populate it with data from the request:
        outForm = ProcedureRequestDownloadForm(request.POST)
        # check whether it's valid:
        if outForm.is_valid():
            outForm_dict = {}
            for item in outForm.cleaned_data.keys():
                if type(outForm.cleaned_data[item]) is not datetime.date:
                    outForm_dict[item] = outForm.cleaned_data[item]
            print "final dict"
            outForm_dict['request_xls_template_path'] = EXCELFILES_FOLDER + "STEP01-ProcedureRequest.xlsx"
            print outForm_dict
            request.session['outForm'] = outForm_dict
            print request.session['outForm']

            ##outForm_dict['excel'] = download_excel(request)
            env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
            template = env.get_template('procedure_request/procedure_request_download.html')
            return HttpResponse(template.render(valData=outForm_dict))
            ##return HttpResponse(render(request,'procedure_request_download.html',valData = outForm_dict ))
    else:
        myForm = ProcedureRequestDownloadForm()
        env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
        template = env.get_template('procedure_request/procedure_request_download_init.html')
        return HttpResponse(template.render(form=myForm))
        ##return render(request,'procedure_request_download_init.html',{'form':myForm})


@csrf_exempt
def download_excel(request):
    session_data = request.session['outForm']
    update_db(session_data)
    xlsx_data = update_excel(session_data)
    ##print xlsx_data
    ##wb = load_workbook(xlsx_data)
    ##wb.save(output)
    out_file = session_data["request_id"] + "_STEP01-ProcedureRequest.xlsx"
    responseDat = HttpResponse(xlsx_data.getvalue(), content_type='application/vnd.ms-excel')
    responseDat['Content-Disposition'] = 'attachment; filename=' + out_file
    ##responseDat.write(xlsx_data)
    xlsx_data.close()
    return responseDat


def update_excel(data):
    print data
    wb = load_workbook(data['request_xls_template_path'])
    ws = wb.active
    ws['D5'] = data['project_name']
    ws['D6'] = data['project_id']
    ws['D7'] = data['project_driver']
    ws['D8'] = data['driver_email']
    ws['D9'] = data['protocol_type']
    ws['D10'] = data['request_id']
    ws['D11'] = data['project_archivist']
    ws['D12'] = data['requested_by']
    ws['D13'] = data['request_date']
    ws['D14'] = data['expected_end_date']
    ws['D15'] = data['quote_number']
    ws['D16'] = data['number_of_samples']
    ws['D17'] = data['number_of_reads']
    ws['D18'] = data['sample_amount']
    ws['D19'] = data['sample_volume']
    ##dest_filename = str(EXCELFILES_FOLDER+data["request_id"]+"_STEP01-ProcedureRequest.xlsx")
    output = StringIO.StringIO()
    ##wb.save(dest_filename)
    ##dest_filename = os.path.join("exceltemplates", data["request_id"]+"_STEP01-ProcedureRequest.xlsx")
    wb.save(output)
    ##return dest_filename
    return output


def update_db(data):
    p = ProjectInfo()
    p.project_id = data['project_id']
    p.project_driver = data['project_driver']
    p.project_name = data['project_name']
    p.driver_email = data['driver_email']
    p.request_id = data['request_id']
    p.expected_completion_date = data['expected_end_date']
    p.save()
    return
