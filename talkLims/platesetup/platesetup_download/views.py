__author__ = "Ashok Ragavendran"
__license__ = "GPL"
__version__ = "1.0.1"
__email__ = "ashok.ragavendran@gmail.com"
__maintainer__ = "Ashok Ragavendran"
__status__ = "Production"

import StringIO
import datetime

from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from jinja2 import FileSystemLoader
from openpyxl import load_workbook

from talkLims.models import ProjectInfo, RequestInfo, PullInfo
from talkLims.platesetup.platesetup_download.platesetupdownload import PlatesetupDownloadForm
from talkLims.settings import EXCELFILES_FOLDER, TEMPLATE_DIRS
from talkLims.utils.jinja2config import environment


@csrf_exempt
def initiate_request(request):
    # if this is a POST request we need to process the form data
    env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
    if request.method == 'POST':
        # create a form instance and populate it with data from the request:
        outForm = PlatesetupDownloadForm(request.POST)
        # check whether it's valid:
        if outForm.is_valid():
            outForm_dict = {}
            for item in outForm.cleaned_data.keys():
                if type(outForm.cleaned_data[item]) is not datetime.date:
                    outForm_dict[item] = outForm.cleaned_data[item]
            print "final dict"
            outForm_dict['mult_requests'] = False
            if len(outForm_dict['request_id'].split('\n')) > 1:
                outForm_dict['mult_requests'] = True
            outForm_dict['request_xls_template_path'] = EXCELFILES_FOLDER + "STEP03-MakePlateforRNASeq_BBC082416.xlsx"
            print outForm_dict
            request.session['outForm'] = outForm_dict
            print request.session['outForm']
            ##db_params(outForm_dict)
            ##outForm_dict['excel'] = download_excel(request)
            env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
            template = env.get_template('plate_setup/plate_setup_download.html')
            return HttpResponse(template.render(valData=outForm_dict))
    else:
        myForm = PlatesetupDownloadForm()
        env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
        template = env.get_template('plate_setup/plate_setup_download_init.html')
        return HttpResponse(template.render(form=myForm))


@csrf_exempt
def download_excel(request):
    session_data = request.session['outForm']
    # update_db(session_data)
    xlsx_data = update_excel(session_data)
    ##print xlsx_data
    ##wb = load_workbook(xlsx_data)
    ##wb.save(output)
    out_file = session_data["request_id"] + "_STEP03-MakePlateforRNASeq_BBC082416.xlsx"
    responseDat = HttpResponse(xlsx_data.getvalue(), content_type='application/vnd.ms-excel')
    responseDat['Content-Disposition'] = 'attachment; filename=' + out_file
    ##responseDat.write(xlsx_data)
    xlsx_data.close()
    return responseDat


def db_proj_params(data):
    if data['mult_requests']:
        print "doing nothing"
        proj_info = []
        req_info = []
        pull_info = []
        for req_id in data['request_id'].split('\n'):
            proj_info.append(ProjectInfo.objects.values().filter(request_id=req_id.strip('\r')))
            req_info.append(RequestInfo.objects.values().filter(request_id=req_id.strip('\r')))
            pull_info.append(PullInfo.objects.values().filter(request_id=req_id.strip('\r')))
    else:
        print data['request_id']
        proj_info = ProjectInfo.objects.values().get(request_id=data['request_id'])
        req_info = RequestInfo.objects.values().get(request_id=data['request_id'])
        pull_info = PullInfo.objects.values().filter(request_id=data['request_id'].strip('\r'))
    print proj_info
    print req_info
    print pull_info
    return proj_info, req_info, pull_info


def update_excel(data):
    proj_info, req_info, pull_info = db_proj_params(data)
    if data['mult_requests']:
        print "do nothing"
    else:

        wb = load_workbook(data['request_xls_template_path'])
        ws = wb.active
        ws['E5'] = proj_info['project_name']
        ws['E6'] = proj_info['project_id']
        ws['E7'] = proj_info['project_driver']
        ws['E8'] = req_info['protocol_type']
        ws['E9'] = req_info['request_id']
        ws['E10'] = req_info['expected_completion_date']
        ws['E11'] = req_info['number_of_samples']

        ws['E12'] = req_info['amount_required']
        ws['E13'] = req_info['request_archivist']
        ws['E14'] = pull_info[0]['pull_date']
        ws['E15'] = data['plate_name']
        ws['E16'] = data['plate_id']
        ws['E17'] = pull_info[0]['assigned_tech']
        sample_num = 1
        start_cell = 34
        heads = ['sample #', 'well', 'barcode_id', 'scan_id', 'sample_name', 'amount_pull', 'vol_pull',
                 'comments_from_request']
        cols = map(chr, range(65, 84))
        well_col = map(chr, range(65, 73))
        well_row = range(1, 13, 1)
        col_cntr = 0
        row_cntr = 1
        form_est_dilution = ''
        form_dil_conc = ''
        form_agilent_corr_ng_per_ul = ''
        form_ul_for_desired_amount = ''
        form_water_to_49_ul = ''

        for i in range(0, len(pull_info)):
            cell_id = []
            for j in range(0, len(cols)):
                cell_id.append(cols[j] + str(start_cell + i))
            ws[cell_id[0]] = i + 1
            ws[cell_id[1]] = well_col[col_cntr] + str(row_cntr).zfill(2)
            col_cntr = col_cntr + 1

            if col_cntr > len(well_col):
                col_cntr = 0
                row_cntr = row_cntr + 1

            ws[cell_id[2]] = pull_info[i]['barcode_id']

            ws[cell_id[4]] = pull_info[i]['sample_name']
            ws[cell_id[5]] = pull_info[i]['amount_from_pull']
            ws[cell_id[6]] = pull_info[i]['vol_from_pull']

            # Form ws[cell_id[7]] = plate_info[i]['est_dilution']
            if i == 0:
                form_est_dilution = ws[cell_id[7]].value

            else:
                ws[cell_id[7]] = form_est_dilution

            ##ws[cell_id[8]] = plate_info[i]['dilution_for_agilent']

            # form ws[cell_id[9]] = plate_info[i]['dil_conc']
            if i == 0:
                form_dil_conc = ws[cell_id[9]].value
            else:
                ws[cell_id[9]] = form_dil_conc

            ##ws[cell_id[10]] = plate_info[i]['rna_rin']
            ##ws[cell_id[11]] = plate_info[i]['ribo_28S_16S']
            ##ws[cell_id[12]] = plate_info[i]['agilent_ng_per_ul']

            # form ws[cell_id[13]] = plate_info[i]['agilent_corr_ng_per_ul']
            if i == 0:
                form_agilent_corr_ng_per_ul = ws[cell_id[13]].value
            else:
                ws[cell_id[13]] = form_agilent_corr_ng_per_ul

            ##ws[cell_id[14]] = plate_info[i]['conc_to_use']
            # formws[cell_id[15]] = plate_info[i]['ul_for_desired_amount']
            if i == 0:
                form_ul_for_desired_amount = ws[cell_id[15]].value
            else:
                ws[cell_id[15]] = form_ul_for_desired_amount

            # formws[cell_id[16]] = plate_info[i]['water_to_49_ul']
            if i == 0:
                form_water_to_49_ul = ws[cell_id[16]].value
            else:
                ws[cell_id[16]] = form_water_to_49_ul
                ##ws[cell_id[17]] = plate_info[i]['ercc']
                ##ws[cell_id[18]] = plate_info[i]['comments']

        output = StringIO.StringIO()

    wb.save(output)
    return output
