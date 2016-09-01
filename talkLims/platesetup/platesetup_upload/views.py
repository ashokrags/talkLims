__author__ = "Ashok Ragavendran"
__license__ = "GPL"
__version__ = "1.0.1"
__email__ = "ashok.ragavendran@gmail.com"
__maintainer__ = "Ashok Ragavendran"
__status__ = "Production"

from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from jinja2 import FileSystemLoader
from openpyxl import load_workbook

from talkLims.models import PlateSetup
from talkLims.platesetup.platesetup_upload.platesetupupload import PlateSetupUploadForm
from talkLims.settings import TEMPLATE_DIRS
from talkLims.utils.jinja2config import environment


@csrf_exempt
def initiate_request(request):
    # if this is a POST request we need to process the form data
    env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
    if request.method == 'POST':
        print request.FILES.keys()
        # create a form instance and populate it with data from the request:
        outForm = PlateSetupUploadForm(request.POST)

        # check whether it's valid:
        print "form loaded"
        if outForm.is_valid():
            print outForm.cleaned_data
        print "form validates"
        print request.FILES['manifest_file']
        update_platessetup_db(request.FILES['manifest_file'])

        env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
        template = env.get_template('base.html')
        return HttpResponse(template.render())
    else:
        myForm = PlateSetupUploadForm()
        env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
        template = env.get_template('plate_setup/plate_setup_upload_init.html')
        return HttpResponse(template.render(form=myForm))


def update_platessetup_db(xls_file):
    wb = load_workbook(xls_file)
    ws = wb.active

    req_id = ws['E9'].value
    plate_name = ws['E15'].value
    plate_id = ws['E16'].value
    plate_made_date = ws['E18'].value
    plate_tech = ws['E19'].value
    agilent_urls = []

    for cell in ws['P5':'P29']:
        print cell
        if cell[0].value is not None:
            agilent_urls.append(cell.value)
        else:
            break

    if len(agilent_urls) > 0:
        agilent_urls = ';'.join(agilent_urls)
    else:
        agilent_urls = 'NONE'

    xls_keys = ['sample_number', 'well', 'barcode_id', 'scan_id', 'sample_name', 'amount_from_plate',
                'vol_from_plate', 'est_dilution', 'dilution_for_agilent', 'dil_conc', 'rna_rin',
                'ribo_28S_16S', 'agilent_ng_per_ul', 'agilent_corr_ng_per_ul', 'conc_to_use',
                'ul_for_desired_amount', 'water_to_49_ul', 'ercc', 'comments']

    for row in ws.iter_rows('A34:K1000'):
        p = PlateSetup()
        ##upload_dat = {k:v for k,v in data.items()}
        upload_dat = {'request_id': req_id, 'plate_id': plate_id, 'plate_name': plate_name,
                      'plate_made_date': plate_made_date, 'plate_tech': plate_tech, 'agilent_location': agilent_urls}
        samples = []
        cell_first = True
        for cell in row:
            val = 0
            ##print cell.value
            if cell.value is not None:
                if isinstance(cell.value, basestring) and cell.value.find('=ROUNDUP') > -1:
                    val = int((float(row[5].value) / 100) + 0.5)
                    ##print cell.value, val
                    samples.append(val)
                elif isinstance(cell.value, basestring) and cell.value.find('=F34/I34') > -1:
                    val = int(row[5].value) / float(row[8].value)
                    ##print cell.value, val
                    samples.append(val)
                elif isinstance(cell.value, basestring) and cell.value.find('=M34*I34') > -1:
                    val = float(row[8].value) * float(row[12].value)
                    ##print cell.value, val
                    samples.append(val)
                elif isinstance(cell.value, basestring) and cell.value.find('P34') > -1:
                    val = 49 - float(row[15].value)
                    ##print cell.value, val
                    samples.append(val)
                else:
                    samples.append(cell.value)
            else:
                cell_first = False
                break
        if not cell_first:
            break
        else:
            sample_dat = dict(zip(xls_keys, samples))
            for k, v in sample_dat.items():
                upload_dat[k] = v
            for (key, value) in upload_dat.items():
                print key, str(value)
                setattr(p, key, value)
            p.save()
    return
