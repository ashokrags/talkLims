__author__ = "Ashok Ragavendran"
__license__ = "GPL"
__version__ = "1.0.1"
__email__ = "ashok.ragavendran@gmail.com"
__maintainer__ = "Ashok Ragavendran"
__status__ = "Production"

from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from jinja2 import FileSystemLoader
from openpyxl import load_workbook

from talkLims.librarymaking.librarymaking_upload.librarymaking_upload import LibraryMakingUploadForm
from talkLims.models import LibraryMaking, LibraryMaking_barcodes
from talkLims.settings import TEMPLATE_DIRS
from talkLims.utils.jinja2config import environment


@csrf_exempt
def initiate_request(request):
    # if this is a POST request we need to process the form data
    env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
    if request.method == 'POST':
        print request.FILES.keys()
        # create a form instance and populate it with data from the request:
        outForm = LibraryMakingUploadForm(request.POST)

        # check whether it's valid:
        print "form loaded"
        if outForm.is_valid():
            print outForm.cleaned_data
        print "form validates"
        print request.FILES['manifest_file']
        update_libmaking_db(request.FILES['manifest_file'])

        env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
        template = env.get_template('base.html')
        return HttpResponse(template.render())
    else:
        myForm = LibraryMakingUploadForm()
        env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
        template = env.get_template('library_making/librarymaking_upload_init.html')
        return HttpResponse(template.render(form=myForm))


def update_libmaking_db(xls_file):
    wb = load_workbook(xls_file)
    ws = wb.active
    upload_dat = dict()
    upload_dat['plate_id'] = ws['E6'].value
    # upload_dat['plate_tech'] = ws['E8'].value
    # upload_dat['plate_made_date'] = ws['E9'].value
    # upload_dat['protocol_id'] = ws['E10'].value
    # upload_dat['number_samples'] = ws['E11'].value

    upload_dat['day1_date'] = ws['O29'].value
    upload_dat['day1_comments'] = ws['P29'].value
    upload_dat['day2_date'] = ws['O112'].value
    upload_dat['day2_comments'] = ws['P112'].value
    upload_dat['final_date'] = ws['O231'].value
    upload_dat['final_comments'] = ws['P231'].value

    upload_dat['add_atl'] = ws['P119'].value
    upload_dat['add_lig'] = ws['P133'].value
    upload_dat['add_smm'] = ws['P89'].value
    upload_dat['add_stl'] = ws['P151'].value
    upload_dat['amp_pcr'] = ws['P214'].value
    upload_dat['clean_up_alp'] = ws['P180'].value
    upload_dat['clean_up_pcr'] = ws['P229'].value
    upload_dat['incubate_1_alp'] = ws['P122'].value
    upload_dat['incubate_1_cdp'] = ws['P80'].value
    upload_dat['incubate_1_rbp'] = ws['P41'].value
    upload_dat['incubate_2_alp'] = ws['P147'].value
    upload_dat['incubate_2_cdp'] = ws['P94'].value
    upload_dat['incubate_2_rbp'] = ws['P63'].value
    upload_dat['incubate_rfp'] = ws['P66'].value
    upload_dat['make_cdp'] = ws['P76'].value
    upload_dat['make_pcr'] = ws['P207'].value
    upload_dat['make_rbp'] = ws['P37'].value
    upload_dat['purify_cdp'] = ws['P109'].value
    upload_dat['run_analytical_pcr'] = ws['P191'].value
    upload_dat['step_1_sample_prep'] = ws['P30'].value
    upload_dat['step_2_synthesize_first_strand_dna_50_min'] = ws['P68'].value
    upload_dat['step_3_synthesize_second_strand_dna_and_clean_up_1_hour'] = ws['P82'].value
    upload_dat['step_4_adenylate_3_pr_ends'] = ws['P113'].value
    upload_dat['step_5_enrich_dna_fragments'] = ws['P182'].value
    upload_dat['wash_rbp'] = ws['P50'].value

    p = LibraryMaking()
    for (key, value) in upload_dat.items():
        print key, str(value)
        if value is None:
            value = 'NULL'
        setattr(p, key, value)

    p.save()
    update_barcode_assignment(ws)
    return


def update_barcode_assignment(xlsx_ws):
    plate_id = xlsx_ws['E6'].value
    print "plate_id:", plate_id
    well_row = map(chr, range(65, 77))
    row_cntr = 0
    col_cntr = 0
    assigned_barcodes = dict()
    for row in xlsx_ws.iter_rows('B136:M143'):
        well_row_pos = well_row[row_cntr]

        for cell in row:
            if cell.value is not None:
                pos = well_row_pos + str(col_cntr + 1).zfill(2)
                assigned_barcodes[pos] = cell.value
            col_cntr = col_cntr + 1
        row_cntr = row_cntr + 1
        col_cntr = 0
    for (key, value) in assigned_barcodes.items():
        p = LibraryMaking_barcodes()
        print key, str(value)
        if value is None:
            value = 'NULL'
        setattr(p, 'barcode_location', key)
        setattr(p, 'plate_id', plate_id)
        setattr(p, 'plate_sample_number', value)
        p.save()
    return
