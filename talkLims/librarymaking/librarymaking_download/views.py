__author__ = "Ashok Ragavendran"
__license__ = "GPL"
__version__ = "1.0.1"
__email__ = "ashok.ragavendran@gmail.com"
__maintainer__ = "Ashok Ragavendran"
__status__ = "Production"

import StringIO
import datetime

from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from jinja2 import FileSystemLoader
from openpyxl import load_workbook

from talkLims.librarymaking.librarymaking_download.librarymaking_download import LibraryMakingDownloadForm
from talkLims.settings import EXCELFILES_FOLDER, TEMPLATE_DIRS
from talkLims.utils.jinja2config import environment


@csrf_exempt
def initiate_request(request):
    # if this is a POST request we need to process the form data
    env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
    if request.method == 'POST':
        # create a form instance and populate it with data from the request:
        outForm = LibraryMakingDownloadForm(request.POST)
        # check whether it's valid:
        if outForm.is_valid():
            outForm_dict = {}
            for item in outForm.cleaned_data.keys():
                if type(outForm.cleaned_data[item]) is not datetime.date:
                    outForm_dict[item] = outForm.cleaned_data[item]
            print "final dict"
            outForm_dict['request_xls_template_path'] = EXCELFILES_FOLDER + "STEP04-RNASeqProcedure_BBC_082416.xlsx"
            print outForm_dict
            request.session['outForm'] = outForm_dict
            print request.session['outForm']

            env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
            template = env.get_template('library_making/librarymaking_download.html')
            return HttpResponse(template.render(valData=outForm_dict))
    else:
        myForm = LibraryMakingDownloadForm()
        env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
        template = env.get_template('library_making/librarymaking_download_init.html')
        return HttpResponse(template.render(form=myForm))


@csrf_exempt
def download_excel(request):
    session_data = request.session['outForm']
    xlsx_data = update_excel(session_data)
    ##print xlsx_data
    ##wb = load_workbook(xlsx_data)
    ##wb.save(output)
    out_file = session_data["plate_id"] + "_STEP04-RNASeqProcedure_BBC_082416.xlsx"
    responseDat = HttpResponse(xlsx_data.getvalue(), content_type='application/vnd.ms-excel')
    responseDat['Content-Disposition'] = 'attachment; filename=' + out_file
    ##responseDat.write(xlsx_data)
    xlsx_data.close()
    return responseDat


@csrf_exempt
def update_excel(data):
    print data
    wb = load_workbook(data['request_xls_template_path'])
    ws = wb.active
    # ws['D5'] =data['project_name']
    # ws['D6'] =data['project_id']
    # ws['D7'] =data['project_driver']
    # ws['D8'] =data['driver_email']
    # ws['D9'] =data['protocol_type']
    ws['E6'] = data['plate_name']
    ws['E7'] = data['plate_id']
    ws['E8'] = data['plate_tech']
    ws['E9'] = data['plate_made_date']
    ##ws['E10'] = data['protocol_id']
    ws['E11'] = data['number_of_samples']

    ##dest_filename = str(EXCELFILES_FOLDER+data["request_id"]+"_STEP01-ProcedureRequest.xlsx")
    output = StringIO.StringIO()
    ##wb.save(dest_filename)
    ##dest_filename = os.path.join("exceltemplates", data["request_id"]+"_STEP01-ProcedureRequest.xlsx")
    wb.save(output)
    ##return dest_filename
    return output
