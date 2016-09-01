__author__ = "Ashok Ragavendran"
__license__ = "GPL"
__version__ = "1.0.1"
__email__ = "ashok.ragavendran@gmail.com"
__maintainer__ = "Ashok Ragavendran"
__status__ = "Production"

from django.core.urlresolvers import reverse
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook

from talkLims.models import PlateSetup
from talkLims.procedurerequest.procedurerequest_upload.procedurerequpload import ProcedureRequestUploadForm
from talkLims.settings import TEMPLATE_DIRS


def environment(**options):
    env = Environment(**options)
    env.globals.update({
        ##'static': staticfiles_storage.url,
        ##'static': "file://Users/ashok//Documents/Research/CHGR/CSSForR/bootstrap-3.3.5-dist/css/bootstrap.min.cyborg.css",
        'static': "https://maxcdn.bootstrapcdn.com/bootswatch/3.3.7/slate/bootstrap.min.css",
        ##'static': "https://maxcdn.bootstrapcdn.com/bootswatch/3.3.7/bootstrap.min.css",
        'url': reverse,
    })
    return env


@csrf_exempt
def initiate_request(request):
    # if this is a POST request we need to process the form data
    env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
    if request.method == 'POST':
        print request.FILES.keys()
        # create a form instance and populate it with data from the request:
        outForm = ProcedureRequestUploadForm(request.POST)

        # check whether it's valid:
        print "form loaded"
        if outForm.is_valid():
            print outForm.cleaned_data
        print "form validates"
        update_platessetup_db(request.FILES['manifest_file'])

        env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
        template = env.get_template('base.html')
        return HttpResponse(template.render())
    else:
        myForm = ProcedureRequestUploadForm()
        env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
        template = env.get_template('procedure_request_upload_init.html')
        return HttpResponse(template.render(form=myForm))


def update_platessetup_db(xls_file):
    wb = load_workbook(xls_file)
    ws = wb.active

    req_id = ws['D9'].value
    plate_name = ws['D15'].value
    plate_id = ws['D16'].value
    plate_made_date = ws['D18'].value
    plate_tech = ws['D19'].value
    agilent_urls = []
    for cell in ws['P5:P29']:
        if cell.value is not None:
            agilent_urls.append(cell.value)
        else:
            break
    agilent_urls = ';'.join(agilent_urls)

    xls_keys = ['sample_number', 'well', 'barcode_id', 'scan_id', 'amount_from_plate',
                'vol_from_plate', 'est_dilution', 'dilution_for_agilent', 'dil_conc', 'rna_rin',
                'ribo_28S_16S', 'agilent_ng_per_ul', 'agilent_corr_ng_per_ul', 'conc_to_use',
                'ul_for_desired_amount', 'water_to_49_ul', 'ercc', 'comments']

    for row in ws.iter_rows('A34:K1000'):
        p = PlateSetup()

        ##upload_dat = {k:v for k,v in data.items()}
        upload_dat = {'request_id': req_id, 'plate_id': plate_id, 'plate_name': plate_name,
                      'plate_made_date': plate_made_date, 'plate_tech': plate_tech, 'agilent_location': agilent_urls}

        samples = []
        cell_first = True
        for cell in row:
            if cell.value is not None:
                samples.append(cell.value)
            else:
                cell_first = False
                break
        if not cell_first:
            break
        else:
            sample_dat = dict(zip(xls_keys, samples))
            for k, v in sample_dat.items():
                upload_dat[k] = v
            for (key, value) in upload_dat.items():
                setattr(p, key, value)
            p.save()
    return
