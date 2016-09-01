__author__ = "Ashok Ragavendran"
__license__ = "GPL"
__version__ = "1.0.1"
__email__ = "ashok.ragavendran@gmail.com"
__maintainer__ = "Ashok Ragavendran"
__status__ = "Production"

import StringIO
import datetime

from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from jinja2 import FileSystemLoader
from openpyxl import load_workbook

from talkLims.models import ProjectInfo, RequestInfo, TempPullRequest
from talkLims.pullrequest.pullrequest_download.pullrequestdownload import PullRequestDownloadForm
from talkLims.settings import EXCELFILES_FOLDER, TEMPLATE_DIRS
from talkLims.utils.jinja2config import environment


@csrf_exempt
def initiate_request(request):
    # if this is a POST request we need to process the form data
    env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
    if request.method == 'POST':
        # create a form instance and populate it with data from the request:
        outForm = PullRequestDownloadForm(request.POST)
        # check whether it's valid:
        if outForm.is_valid():
            outForm_dict = {}
            for item in outForm.cleaned_data.keys():
                if type(outForm.cleaned_data[item]) is not datetime.date:
                    outForm_dict[item] = outForm.cleaned_data[item]
            print "final dict"
            outForm_dict['request_xls_template_path'] = EXCELFILES_FOLDER + "STEP02-SamplePull_BBC_082416.xlsx"
            print outForm_dict
            request.session['outForm'] = outForm_dict
            print request.session['outForm']
            db_params(outForm_dict)
            env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
            template = env.get_template('pull_request/pull_request_download.html')
            return HttpResponse(template.render(valData=outForm_dict))
    else:
        myForm = PullRequestDownloadForm()
        env = environment(loader=FileSystemLoader(TEMPLATE_DIRS))
        template = env.get_template('pull_request/pull_request_download_init.html')
        return HttpResponse(template.render(form=myForm))


@csrf_exempt
def download_excel(request):
    session_data = request.session['outForm']
    xlsx_data = update_excel(session_data)
    ##print xlsx_data
    ##wb = load_workbook(xlsx_data)
    ##wb.save(output)
    out_file = session_data["request_id"] + "_STEP02-SamplePull_BBC_082416.xlsx"
    responseDat = HttpResponse(xlsx_data.getvalue(), content_type='application/vnd.ms-excel')
    responseDat['Content-Disposition'] = 'attachment; filename=' + out_file
    ##responseDat.write(xlsx_data)
    xlsx_data.close()
    return responseDat


def db_params(data):
    project_info = ProjectInfo.objects.filter(request_id=data['request_id']).values()
    print project_info
    req_info = RequestInfo.objects.filter(request_id=data['request_id']).values()
    print req_info
    tempDat = TempPullRequest.objects.filter(request_id=data['request_id']).values()
    print tempDat
    return


def update_excel(data):
    project_info = ProjectInfo.objects.filter(request_id=data['request_id']).values()
    req_info = RequestInfo.objects.filter(request_id=data['request_id']).values()
    tempDat = TempPullRequest.objects.filter(request_id=data['request_id']).values()
    wb = load_workbook(data['request_xls_template_path'])
    ws = wb.active
    ws['D5'] = project_info[0]['project_name']
    ws['D6'] = project_info[0]['project_id']
    ws['D7'] = project_info[0]['project_driver']
    ws['D8'] = req_info[0]['protocol_type']
    ws['D9'] = req_info[0]['request_id']
    ws['D10'] = req_info[0]['requested_by']
    ws['D11'] = req_info[0]['request_date']
    ws['D12'] = project_info[0]['expected_completion_date']
    ws['D13'] = req_info[0]['quote_number']
    ws['D14'] = req_info[0]['number_of_samples']
    ws['D15'] = req_info[0]['number_of_reads']
    ws['D16'] = req_info[0]['amount_required']
    ws['D17'] = req_info[0]['volume_required']
    ws['D18'] = req_info[0]['request_archivist']
    # ws['D19'] = RequestInfo[0][]
    ##dest_filename = str(EXCELFILES_FOLDER+data["request_id"]+"_STEP01-ProcedureRequest.xlsx")
    start_cell = 34
    heads = ['scan_id', 'sample_name', 'box_name', 'location_jane', 'shelf_jane', 'rack_jane',
             'slot_jane', 'grid', 'amount_jane', 'vol_jane', 'comments_from_request']
    cols = map(chr, range(65, 76))
    for i in range(0, len(tempDat)):
        for j in range(0, len(heads)):
            cell_id = cols[j] + str(start_cell + i)
            ws[cell_id] = tempDat[i][heads[j]]

    output = StringIO.StringIO()
    ##wb.save(dest_filename)
    ##dest_filename = os.path.join("exceltemplates", data["request_id"]+"_STEP01-ProcedureRequest.xlsx")
    wb.save(output)
    ##return dest_filename
    return output
