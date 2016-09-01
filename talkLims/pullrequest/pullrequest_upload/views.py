__author__ = "Ashok Ragavendran"
__license__ = "GPL"
__version__ = "1.0.1"
__email__ = "ashok.ragavendran@gmail.com"
__maintainer__ = "Ashok Ragavendran"
__status__ = "Production"

from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from jinja2 import FileSystemLoader
from openpyxl import load_workbook

from talkLims.models import PullInfo
from talkLims.pullrequest.pullrequest_upload.pullrequpload import PullRequestUploadForm
from talkLims.utils.jinja2config import environment


@csrf_exempt
def initiate_request(request):
    # if this is a POST request we need to process the form data
    env = environment(loader=FileSystemLoader('/Users/ashok/PycharmProjects/talkLims/templates'))
    if request.method == 'POST':
        print request.FILES.keys()
        # create a form instance and populate it with data from the request:
        outForm = PullRequestUploadForm(request.POST)

        # check whether it's valid:
        print "form loaded"
        if outForm.is_valid():
            print outForm.cleaned_data
        print "form validates"
        update_db_pull_info(request.FILES['manifest_file'])

        env = environment(loader=FileSystemLoader('/Users/ashok/PycharmProjects/talkLims/templates'))
        template = env.get_template('base.html')
        return HttpResponse(template.render())
    else:
        myForm = PullRequestUploadForm()
        env = environment(loader=FileSystemLoader('/Users/ashok/PycharmProjects/talkLims/templates'))
        template = env.get_template('pull_request/pull_request_upload_init.html')
        return HttpResponse(template.render(form=myForm))


def update_db_pull_info(xls_file):
    wb = load_workbook(xls_file)
    ws = wb.active

    req_id = ws['D9'].value
    pull_date = ws['D19'].value
    assigned_tech = ws['D20'].value

    r1 = [x.values()[0] for x in PullInfo.objects.values('request_id')]

    if len(r1) > 0 and req_id in r1:
        print "Request ID exists in temp Table"
        return
    else:
        print "uploading Table"

    xls_keys_jane = ['barcode_id', 'sample_name', 'location_jane', 'box_name', 'shelf_jane',
                     'rack_jane', 'slot_jane', 'grid', 'amount_jane', 'vol_jane', 'comments_from_request']

    xls_keys_pull = ['scan_id', 'location_from_pull', 'box_name_from_pull', 'shelf_from_pull', 'rack_from_pull',
                     'slot_from_pull', 'grid_from_pull', 'amount_from_pull', 'vol_from_pull', 'comments_for_pull']
    xls_keys = xls_keys_jane + xls_keys_pull
    print xls_keys
    for row in ws.iter_rows('A34:U1000'):
        p = PullInfo()
        ##upload_dat = {k:v for k,v in data.items()}
        upload_dat = {'request_id': req_id, 'pull_date': pull_date, 'assigned_tech': assigned_tech}

        samples = []
        cell_first = True
        for cell in row:
            if cell.value is not None:
                samples.append(cell.value)
            else:
                cell_first = False
                break
        if not cell_first:
            break
        else:
            sample_dat = dict(zip(xls_keys, samples))
            print sample_dat
            for k, v in sample_dat.items():
                upload_dat[k] = v
            print upload_dat
            for (key, value) in upload_dat.items():
                setattr(p, key, value)
            p.save()
    return
